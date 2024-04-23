#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl.utils
import openpyxl.styles
import openpyxl.drawing.image


def gen_text_pos(row, col):
    return "{col}{row}".format(
        row=row,
        col=openpyxl.utils.get_column_letter(col),
    )


def set_header_cell_style(sheet, row, col, value, width, style):
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).fill = style["fill"]

    if width is not None:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width


def insert_table_header(sheet, row, sheet_def, base_style):
    for key in sheet_def["TABLE_HEADER"]["col"].keys():
        col = sheet_def["TABLE_HEADER"]["col"][key]["pos"]
        if "width" in sheet_def["TABLE_HEADER"]["col"][key]:
            width = sheet_def["TABLE_HEADER"]["col"][key]["width"]
        else:
            width = None

        if key == "category":
            for i in range(sheet_def["TABLE_HEADER"]["col"][key]["length"]):
                set_header_cell_style(
                    sheet,
                    row,
                    col + i,
                    sheet_def["TABLE_HEADER"]["col"][key]["label"] + " ({i})".format(i=i + 1),
                    width,
                    base_style,
                )
        else:
            set_header_cell_style(
                sheet, row, col, sheet_def["TABLE_HEADER"]["col"][key]["label"], width, base_style
            )


def gen_item_cell_style(base_style, cell_def):
    style = base_style.copy()

    if "format" in cell_def:
        style["text_format"] = cell_def["format"]

    if "wrap" in cell_def:
        style["text_wrap"] = cell_def["wrap"]
    else:
        style["text_wrap"] = False

    return style


def set_item_cell_style(sheet, row, col, value, style):
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).alignment = openpyxl.styles.Alignment(wrap_text=style["text_wrap"], vertical="top")

    if "text_format" in style:
        sheet.cell(row, col).number_format = style["text_format"]


def insert_table_item(sheet, row, item, is_need_thumb, thumb_path, sheet_def, base_style):
    for key in sheet_def["TABLE_HEADER"]["col"].keys():
        col = sheet_def["TABLE_HEADER"]["col"][key]["pos"]

        cell_style = gen_item_cell_style(base_style, sheet_def["TABLE_HEADER"]["col"][key])

        if key == "category":
            for i in range(sheet_def["TABLE_HEADER"]["col"][key]["length"]):
                if i < len(item["category"]):
                    value = item[key][i]
                else:
                    value = ""
                set_item_cell_style(sheet, row, col + i, value, cell_style)
        elif key == "image":
            sheet.cell(row, col).border = cell_style["border"]
            if is_need_thumb:
                insert_table_cell_image(
                    sheet,
                    row,
                    col,
                    thumb_path,
                    sheet_def["TABLE_HEADER"]["col"]["image"]["width"],
                    sheet_def["TABLE_HEADER"]["row"]["height"],
                )
        else:
            if (
                ("optional" in sheet_def["TABLE_HEADER"]["col"][key])
                and sheet_def["TABLE_HEADER"]["col"][key]["optional"]
                and (key not in item)
            ):
                value = None
            else:
                if "value" in sheet_def["TABLE_HEADER"]["col"][key]:
                    value = sheet_def["TABLE_HEADER"]["col"][key]["value"]
                elif "formal_key" in sheet_def["TABLE_HEADER"]["col"][key]:
                    value = item[sheet_def["TABLE_HEADER"]["col"][key]["formal_key"]]
                else:
                    value = item[key]

                if "conv_func" in sheet_def["TABLE_HEADER"]["col"][key]:
                    value = sheet_def["TABLE_HEADER"]["col"][key]["conv_func"](value)

            set_item_cell_style(sheet, row, col, value, cell_style)

        if "link_func" in sheet_def["TABLE_HEADER"]["col"][key]:
            sheet.cell(row, col).hyperlink = sheet_def["TABLE_HEADER"]["col"][key]["link_func"](item)


def insert_table_cell_image(sheet, row, col, thumb_path, cell_width, cell_height):
    if (thumb_path is None) or (not thumb_path.exists()):
        return

    img = openpyxl.drawing.image.Image(thumb_path)

    # NOTE: マジックナンバー「8」は下記等を参考にして設定．(日本語フォントだと 8 が良さそう)
    # > In all honesty, I cannot tell you how many blogs and stack overflow answers
    # > I read before I stumbled across this magic number: 7.5
    # https://imranhugo.medium.com/how-to-right-align-an-image-in-excel-cell-using-python-and-openpyxl-7ca75a85b13a
    cell_width_pix = cell_width * 8
    cell_height_pix = openpyxl.utils.units.points_to_pixels(cell_height)

    cell_width_emu = openpyxl.utils.units.pixels_to_EMU(cell_width_pix)
    cell_height_emu = openpyxl.utils.units.pixels_to_EMU(cell_height_pix)

    margin_pix = 2
    content_width_pix = cell_width_pix - (margin_pix * 2)
    content_height_pix = cell_height_pix - (margin_pix * 2)

    content_ratio = content_width_pix / content_height_pix
    image_ratio = img.width / img.height

    if (img.width > content_width_pix) or (img.height > content_height_pix):
        if image_ratio > content_ratio:
            # NOTE: 画像の横幅をセルの横幅に合わせる
            scale = content_width_pix / img.width
        else:
            # NOTE: 画像の高さをセルの高さに合わせる
            scale = content_height_pix / img.height

        img.width *= scale
        img.height *= scale

    image_width_emu = openpyxl.utils.units.pixels_to_EMU(img.width)
    image_height_emu = openpyxl.utils.units.pixels_to_EMU(img.height)

    col_offset_emu = (cell_width_emu - image_width_emu) / 2
    row_offset_emu = (cell_height_emu - image_height_emu) / 2

    marker_1 = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col - 1, row=row - 1, colOff=col_offset_emu, rowOff=row_offset_emu
    )
    marker_2 = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col, row=row, colOff=-col_offset_emu, rowOff=-row_offset_emu
    )

    img.anchor = openpyxl.drawing.spreadsheet_drawing.TwoCellAnchor(_from=marker_1, to=marker_2)

    sheet.add_image(img)


def setting_table_view(sheet, sheet_def, row_last, is_hidden):
    sheet.column_dimensions.group(
        openpyxl.utils.get_column_letter(sheet_def["TABLE_HEADER"]["col"]["image"]["pos"]),
        openpyxl.utils.get_column_letter(sheet_def["TABLE_HEADER"]["col"]["image"]["pos"]),
        hidden=is_hidden,
    )

    sheet.freeze_panes = gen_text_pos(
        sheet_def["TABLE_HEADER"]["row"]["pos"] + 1,
        sheet_def["TABLE_HEADER"]["col"]["price"]["pos"] + 1,
    )

    sheet.auto_filter.ref = "{start}:{end}".format(
        start=gen_text_pos(
            sheet_def["TABLE_HEADER"]["row"]["pos"],
            min(map(lambda x: x["pos"], sheet_def["TABLE_HEADER"]["col"].values())),
        ),
        end=gen_text_pos(row_last, max(map(lambda x: x["pos"], sheet_def["TABLE_HEADER"]["col"].values()))),
    )
    sheet.sheet_view.showGridLines = False


def generate_list_sheet(
    book,
    item_list,
    sheet_def,
    is_need_thumb,
    thumb_path_func,
    set_status_func,
    update_seq_func,
    update_item_func,
):
    sheet = book.create_sheet()
    sheet.title = "{label}アイテム一覧".format(label=sheet_def["SHEET_TITLE"])

    side = openpyxl.styles.Side(border_style="thin", color="000000")
    border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="F2F2F2")

    base_style = {"border": border, "fill": fill}

    row = sheet_def["TABLE_HEADER"]["row"]["pos"]

    set_status_func("テーブルのヘッダを設定しています...")
    insert_table_header(sheet, row, sheet_def, base_style)

    update_seq_func()

    set_status_func("{label} - 商品の記載をしています...".format(label=sheet_def["SHEET_TITLE"]))

    row += 1
    for item in item_list:
        sheet.row_dimensions[row].height = sheet_def["TABLE_HEADER"]["row"]["height"]
        insert_table_item(sheet, row, item, is_need_thumb, thumb_path_func(item), sheet_def, base_style)
        update_item_func()

        row += 1

    row_last = row - 1

    update_item_func()
    update_seq_func()

    set_status_func("テーブルの表示設定しています...")
    setting_table_view(sheet, sheet_def, row_last, not is_need_thumb)

    update_seq_func()
